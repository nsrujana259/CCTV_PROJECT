"""
GVPCEW HOSTEL SURVEILLANCE — FAST VERSION
- Haar cascade: runs every frame (microseconds) = smooth video
- ArcFace: runs ONCE when a new face appears = identify who they are
- Appear in frame = IN, Disappear = OUT
"""
import warnings
warnings.filterwarnings("ignore")
import streamlit as st
import cv2, os, time, pickle
import threading, queue
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="GVPCEW Surveillance", page_icon="🛡️",
                   layout="wide", initial_sidebar_state="expanded")

ROLL_NUMBERS = {
    'sushmitha' :'322103282070','praneetha':'322103282101',
    'anusha'    :'322103282075','devika'   :'322103282116',
    'navya'     :'322103282118','jessy'    :'322103282065',
    'ashwitha'  :'322103282093','anuradha' :'322103282100',
    'jayasri'   :'322103282096','gayatri'  :'322103282024',
    'vaghdevi'  :'322103282073','srujana'  :'322103282069',
    'jyothsna'  :'322103282114',
}
EXCEL_MASTER  = 'attendance_master.xlsx'   # all sessions, never deleted
# session file path is set at START time (see session state below)

for k,v in [('emb_data',None),('emb_names',[]),('running',False),
             ('stop_flag',False),('in_count',0),('out_count',0),('events',[]),
             ('session_file',None)]:
    if k not in st.session_state: st.session_state[k] = v

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Sans:wght@400;500;600&family=JetBrains+Mono:wght@400;600&display=swap');
html,body,[class*="css"],.stApp{background:#080e1a!important;color:#f0f6fc!important;font-family:'DM Sans',sans-serif!important}
::-webkit-scrollbar{width:5px}::-webkit-scrollbar-thumb{background:#1e3a5f;border-radius:3px}
.hdr{display:flex;align-items:center;gap:18px;padding:20px 26px;background:linear-gradient(135deg,#0c1526,#111d33);border:1px solid rgba(56,189,248,0.25);border-top:3px solid #38bdf8;border-radius:12px;margin-bottom:18px;}
.hdr-icon{width:50px;height:50px;background:linear-gradient(135deg,#1e3a5f,#2563a8);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:22px;flex-shrink:0;}
.hdr h1{font-family:'Bebas Neue',sans-serif;font-size:1.7rem;letter-spacing:3px;line-height:1;margin:0;}
.hdr p{color:#c0cfe0;font-size:0.72rem;margin-top:3px;}
.hdr-badge{margin-left:auto;text-align:right;}
.dot{display:inline-flex;align-items:center;gap:6px;font-family:'JetBrains Mono',monospace;font-size:0.66rem;letter-spacing:1px;}
.dot.on{color:#10b981}.dot.on::before{content:'';width:7px;height:7px;background:#10b981;border-radius:50%;box-shadow:0 0 6px #10b981;animation:p 1.5s infinite;display:inline-block;}
.dot.off{color:#f59e0b}.dot.off::before{content:'';width:7px;height:7px;background:#f59e0b;border-radius:50%;display:inline-block;}
@keyframes p{0%,100%{opacity:1}50%{opacity:0.3}}
.sid{font-family:'JetBrains Mono',monospace;font-size:0.6rem;color:#8baec8;margin-top:2px;}
.mrow{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:16px;}
.mc{background:#0c1526;border:1px solid rgba(56,189,248,0.12);border-radius:10px;padding:16px 18px;position:relative;}
.mc::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;}
.mc.in::before{background:#10b981}.mc.out::before{background:#ef4444}.mc.now::before{background:#38bdf8}
.ml{font-size:0.62rem;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#8baec8;margin-bottom:5px;}
.mv{font-family:'Bebas Neue',sans-serif;font-size:2.6rem;line-height:1;letter-spacing:2px;}
.mc.in .mv{color:#10b981}.mc.out .mv{color:#ef4444}.mc.now .mv{color:#38bdf8}
.ms{font-size:0.65rem;color:#8baec8;margin-top:3px;}
.sh{display:flex;align-items:center;gap:10px;margin-bottom:10px;}
.sh h3{font-family:'Bebas Neue',sans-serif;font-size:0.9rem;letter-spacing:3px;color:#e0eaf5;white-space:nowrap;}
.sh::after{content:'';flex:1;height:1px;background:rgba(56,189,248,0.12);}
.cc{background:#0c1526;border:1px solid rgba(56,189,248,0.12);border-radius:10px;padding:14px;margin-bottom:10px;}
.cc h4{font-family:'Bebas Neue',sans-serif;letter-spacing:2px;font-size:0.82rem;color:#38bdf8;margin-bottom:8px;}
.ip{background:#0c1526;border:1px solid rgba(56,189,248,0.12);border-left:3px solid #38bdf8;border-radius:10px;padding:14px;}
.ip h4{font-family:'Bebas Neue',sans-serif;letter-spacing:2px;color:#38bdf8;margin-bottom:10px;font-size:0.85rem;}
.ir{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid rgba(56,189,248,0.08);font-size:0.76rem;}
.ir:last-child{border-bottom:none;}.ik{color:#9ab5cc;}.iv{color:#f0f6fc;font-family:'JetBrains Mono',monospace;font-size:0.7rem;}
.ab{background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.3);border-left:3px solid #f59e0b;border-radius:7px;padding:9px 12px;font-size:0.76rem;color:#f59e0b;margin:7px 0;}
.sb{background:rgba(16,185,129,0.08);border:1px solid rgba(16,185,129,0.25);border-left:3px solid #10b981;border-radius:7px;padding:9px 12px;font-size:0.76rem;color:#10b981;margin:7px 0;}
.lb{background:rgba(16,185,129,0.05);border:1px solid rgba(16,185,129,0.2);border-left:3px solid #10b981;border-radius:7px;padding:9px 12px;font-size:0.76rem;color:#10b981;margin-bottom:8px;}
.el{background:#0c1526;border:1px solid rgba(56,189,248,0.12);border-radius:10px;overflow:hidden;}
.eh{display:grid;padding:9px 14px;background:#111d33;border-bottom:1px solid rgba(56,189,248,0.12);font-size:0.62rem;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#8baec8;}
.er{display:grid;padding:8px 14px;border-bottom:1px solid rgba(56,189,248,0.08);font-size:0.78rem;align-items:center;}
.er:last-child{border-bottom:none;}
.tin{display:inline-block;background:rgba(16,185,129,0.15);color:#10b981;border:1px solid rgba(16,185,129,0.3);padding:1px 7px;border-radius:4px;font-size:0.66rem;font-weight:700;}
.tout{display:inline-block;background:rgba(239,68,68,0.15);color:#ef4444;border:1px solid rgba(239,68,68,0.3);padding:1px 7px;border-radius:4px;font-size:0.66rem;font-weight:700;}
section[data-testid="stSidebar"]{background:#0c1526!important;border-right:1px solid rgba(56,189,248,0.12)!important;}
div[data-testid="stFileUploader"]>div{background:#0c1526!important;border:2px dashed rgba(56,189,248,0.25)!important;border-radius:10px!important;}
div[data-testid="stButton"] button{width:100%!important;font-family:'Bebas Neue',sans-serif!important;font-size:1.05rem!important;letter-spacing:3px!important;border:none!important;border-radius:8px!important;padding:13px!important;cursor:pointer!important;}
div[data-testid="stButton"]:first-of-type button:not(:disabled){background:#059669!important;color:#ffffff!important;box-shadow:0 4px 14px rgba(5,150,105,0.45)!important;}
div[data-testid="stButton"]:last-of-type button:not(:disabled){background:#dc2626!important;color:#ffffff!important;box-shadow:0 4px 14px rgba(220,38,38,0.45)!important;}
div[data-testid="stButton"] button:disabled{background:#162040!important;color:#4a6480!important;box-shadow:none!important;}
.stDownloadButton>button{background:#111d33!important;color:#38bdf8!important;border:1px solid rgba(56,189,248,0.25)!important;border-radius:8px!important;font-size:0.82rem!important;font-weight:600!important;padding:8px 16px!important;}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding-top:14px!important;padding-bottom:40px!important;}
</style>""", unsafe_allow_html=True)


# ── Helpers ───────────────────────────────────────────────────
def _make_excel_headers(ws):
    headers=['S.No','Name','Roll Number','Date','Time In','Time Out','Duration']
    widths  =[  6,    18,      18,         14,    12,       12,        12     ]
    hf=PatternFill("solid",fgColor="1E3A5F")
    b=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(row=1,column=i,value=h); c.fill=hf
        c.font=Font(bold=True,color="FFFFFF",size=11)
        c.alignment=Alignment(horizontal="center"); c.border=b
        ws.column_dimensions[c.column_letter].width=w

def init_excel():
    # Session file: always fresh, named with timestamp
    session_file=f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.session_state.session_file=session_file
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Session Log"
    _make_excel_headers(ws); wb.save(session_file)
    # Master file: create only if it doesn't exist
    if not os.path.exists(EXCEL_MASTER):
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="All Sessions"
        _make_excel_headers(ws); wb.save(EXCEL_MASTER)

# row_map[filepath][name] = row_number of the open IN entry
_row_map = {}

def _style_row(ws, rn, name, has_out):
    b=Border(left=Side(style='thin'),right=Side(style='thin'),
             top=Side(style='thin'),bottom=Side(style='thin'))
    if name=='Unknown': fc,tc="F0F0F0","888888"
    elif has_out:       fc,tc="E6F9F0","065F46"   # green = complete
    else:               fc,tc="FFF9E6","78550A"   # amber = still inside
    rf=PatternFill("solid",fgColor=fc)
    for col in range(1,8):
        c=ws.cell(row=rn,column=col)
        c.fill=rf; c.border=b
        c.font=Font(color=tc,size=10)
        c.alignment=Alignment(horizontal="center")

def _log_in(filepath, name, roll, now):
    global _row_map
    wb=openpyxl.load_workbook(filepath); ws=wb.active
    rn=ws.max_row+1; sno=rn-1
    ws.cell(rn,1,sno); ws.cell(rn,2,name.title()); ws.cell(rn,3,roll)
    ws.cell(rn,4,now.strftime('%d-%m-%Y')); ws.cell(rn,5,now.strftime('%H:%M:%S'))
    ws.cell(rn,6,'—'); ws.cell(rn,7,'—')
    _style_row(ws,rn,name,False)
    wb.save(filepath)
    if filepath not in _row_map: _row_map[filepath]={}
    _row_map[filepath][name]=rn

def _log_out(filepath, name, now):
    global _row_map
    if filepath not in _row_map or name not in _row_map[filepath]:
        return  # no matching IN row — skip
    rn=_row_map[filepath][name]
    wb=openpyxl.load_workbook(filepath); ws=wb.active
    time_in_str=ws.cell(rn,5).value
    ws.cell(rn,6,now.strftime('%H:%M:%S'))
    # Calculate duration
    try:
        t_in=datetime.strptime(time_in_str,'%H:%M:%S')
        t_out=datetime.strptime(now.strftime('%H:%M:%S'),'%H:%M:%S')
        diff=t_out-t_in
        total_s=int(diff.total_seconds())
        dur=f"{total_s//3600:02d}:{(total_s%3600)//60:02d}:{total_s%60:02d}"
        ws.cell(rn,7,dur)
    except Exception:
        ws.cell(rn,7,'—')
    _style_row(ws,rn,name,True)
    wb.save(filepath)
    del _row_map[filepath][name]  # row is now closed

def log_excel(name,roll,direction,conf):
    try:
        now=datetime.now()
        files=[f for f in [st.session_state.get('session_file'),EXCEL_MASTER] if f]
        for fp in files:
            if direction=='IN':  _log_in(fp,name,roll,now)
            else:                _log_out(fp,name,now)
    except Exception as e: print(f"Excel:{e}")

def load_embeddings(data):
    db={}
    for name,emb in data.items():
        arr=np.array(emb,dtype=np.float32).squeeze()
        if arr.ndim==2:
            arr=arr/(np.linalg.norm(arr,axis=1,keepdims=True)+1e-8); arr=arr.mean(axis=0)
        arr=arr/(np.linalg.norm(arr)+1e-8); db[name]=arr
    return list(db.keys()),np.stack(list(db.values()))

def recognize(emb,names,matrix,threshold):
    e=emb/(np.linalg.norm(emb)+1e-8); sims=matrix@e
    idx=int(np.argmax(sims)); score=float(sims[idx])
    return (names[idx],score) if score>=threshold else ('Unknown',score)


# ── UI ────────────────────────────────────────────────────────
is_live=st.session_state.running; emb_ok=st.session_state.emb_data is not None

st.markdown(f"""<div class="hdr">
  <div class="hdr-icon">🛡️</div>
  <div><h1>HOSTEL SURVEILLANCE SYSTEM</h1>
    <p>GAYATRI VIDYA PARISHAD COLLEGE OF ENGINEERING FOR WOMEN · CSE (AI &amp; ML)</p>
    <p style="color:#8fafc6;font-size:0.68rem;margin-top:1px;">Fast Mode · Appear = IN · Disappear = OUT</p>
  </div>
  <div class="hdr-badge">
    <div class="dot {'on' if is_live else 'off'}">{'● LIVE' if is_live else '○ STANDBY'}</div>
    <div class="sid">GVPCEW-HST-01 · Haar+ArcFace Hybrid</div>
  </div>
</div>""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("""<div style="padding:4px 0 12px;"><div style="font-family:'Bebas Neue',sans-serif;letter-spacing:3px;font-size:0.95rem;color:#38bdf8;">⚙ CONTROL PANEL</div><div style="height:1px;background:rgba(56,189,248,0.15);margin-top:6px;"></div></div>""", unsafe_allow_html=True)

    st.markdown('<div class="cc"><h4>📁 FACE DATABASE</h4>', unsafe_allow_html=True)
    emb_file=st.file_uploader("emb",type=['pkl'],label_visibility="collapsed")
    if emb_file:
        if st.session_state.emb_data is None:
            try:
                st.session_state.emb_data=pickle.load(emb_file)
                st.session_state.emb_names=list(st.session_state.emb_data.keys())
                st.rerun()   # immediately re-render so START button enables
            except Exception as e:
                st.markdown(f'<div class="ab">✗ {e}</div>',unsafe_allow_html=True)
    # Only show warning if nothing loaded yet — don't clear on widget rerun
    if st.session_state.emb_data:
        n=st.session_state.emb_names
        st.markdown(f'<div class="sb">✓ {len(n)} students enrolled</div>',unsafe_allow_html=True)
    else:
        st.markdown('<div class="ab">⚠ Upload arcface_embeddings.pkl</div>',unsafe_allow_html=True)
    st.markdown('</div>',unsafe_allow_html=True)

    st.markdown('<div class="cc"><h4>📷 CAMERA</h4>',unsafe_allow_html=True)
    cam_index=st.selectbox("cam",[0,1,2],format_func=lambda x:f"Camera {x}"+(" (built-in)" if x==0 else " (USB)"),label_visibility="collapsed")
    rtsp_url=st.text_input("RTSP",placeholder="rtsp://admin:pass@192.168.1.64:554/stream",label_visibility="collapsed")
    st.markdown('</div>',unsafe_allow_html=True)

    st.markdown('<div class="cc"><h4>🎯 RECOGNITION</h4>',unsafe_allow_html=True)
    threshold=st.slider("thresh",0.25,0.65,0.32,0.01,label_visibility="collapsed")
    mode='STRICT' if threshold>0.50 else 'BALANCED' if threshold>0.34 else 'LENIENT'
    st.markdown(f'<div style="font-family:JetBrains Mono,monospace;font-size:0.68rem;color:#9ab8ce;">{threshold:.2f} · {mode}</div>',unsafe_allow_html=True)
    st.markdown('</div>',unsafe_allow_html=True)

    st.markdown('<div class="cc"><h4>⏱ COOLDOWN</h4>',unsafe_allow_html=True)
    cooldown=st.slider("cool",3,20,8,label_visibility="collapsed")
    st.markdown(f'<div style="font-family:JetBrains Mono,monospace;font-size:0.68rem;color:#9ab8ce;">{cooldown}s before same person re-logged</div>',unsafe_allow_html=True)
    st.markdown('</div>',unsafe_allow_html=True)

    st.markdown("""<div style="background:#080e1a;border:1px solid rgba(56,189,248,0.08);border-radius:8px;padding:12px;font-size:0.72rem;color:#8fafc6;line-height:2;">
      👤 <b style="color:#10b981">Appear in frame</b> = IN logged<br>
      🚪 <b style="color:#ef4444">Leave frame</b> = OUT logged<br>
      ⚡ Video is smooth — recognition runs once per new face<br>
      📊 Excel logs every event
    </div>""",unsafe_allow_html=True)

col_main,col_side=st.columns([3,2],gap="large")
with col_side:
    st.markdown('<div class="sh"><h3>📊 STATUS</h3></div>',unsafe_allow_html=True)
    inside=max(0,st.session_state.in_count-st.session_state.out_count)
    cam_src="RTSP" if rtsp_url else f"Camera {cam_index}"
    st.markdown(f"""<div class="ip"><h4>🔍 SYSTEM</h4>
      <div class="ir"><span class="ik">Database</span><span class="iv" style="color:{'#10b981' if emb_ok else '#ef4444'};">{'✓ LOADED' if emb_ok else '✗ MISSING'}</span></div>
      <div class="ir"><span class="ik">Camera</span><span class="iv" style="color:{'#10b981' if is_live else '#f59e0b'};">{'✓ LIVE' if is_live else '○ STANDBY'}</span></div>
      <div class="ir"><span class="ik">Source</span><span class="iv">{cam_src}</span></div>
      <div class="ir"><span class="ik">Enrolled</span><span class="iv">{len(st.session_state.emb_names) if emb_ok else '—'}</span></div>
      <div class="ir"><span class="ik">Threshold</span><span class="iv">{threshold:.2f} · {mode}</span></div>
      <div class="ir"><span class="ik">IN</span><span class="iv" style="color:#10b981;">{st.session_state.in_count}</span></div>
      <div class="ir"><span class="ik">OUT</span><span class="iv" style="color:#ef4444;">{st.session_state.out_count}</span></div>
      <div class="ir"><span class="ik">Inside</span><span class="iv" style="color:#38bdf8;">{inside}</span></div>
      <div class="ir"><span class="ik">Events</span><span class="iv">{len(st.session_state.events)}</span></div>
    </div>""",unsafe_allow_html=True)

    if st.session_state.events:
        st.markdown("<br>",unsafe_allow_html=True)
        st.markdown('<div class="sh"><h3>⚡ RECENT</h3></div>',unsafe_allow_html=True)
        ticker=""
        for ev in reversed(st.session_state.events[-6:]):
            n=ev['name'].title() if ev['name']!='Unknown' else 'Unknown'
            inside=ev['out_time']=='—'
            sc="#f59e0b" if inside else "#10b981"
            status="INSIDE" if inside else "LEFT"
            ticker+=(f'<div style="display:flex;align-items:center;gap:9px;padding:7px 11px;border-bottom:1px solid rgba(56,189,248,0.08);font-size:0.74rem;">'
                     f'<div style="flex:1;"><div style="font-weight:500;">{n}</div>'
                     f'<div style="font-family:JetBrains Mono,monospace;font-size:0.63rem;color:#8baec8;">'
                     f'{ev.get("roll","—")} · IN {ev["in_time"]}'
                     f'{" · OUT "+ev["out_time"] if not inside else ""}'
                     f'</div></div>'
                     f'<span style="color:{sc};font-size:0.64rem;font-weight:700;">{status}</span></div>')
        st.markdown(f'<div style="background:#0c1526;border:1px solid rgba(56,189,248,0.12);border-radius:10px;overflow:hidden;">{ticker}</div>',unsafe_allow_html=True)

with col_main:
    st.markdown('<div class="sh"><h3>📹 LIVE FEED</h3></div>',unsafe_allow_html=True)
    feed_ph=st.empty()
    if not is_live:
        feed_ph.markdown("""<div style="background:#0c1526;border:2px dashed rgba(56,189,248,0.18);border-radius:10px;padding:70px 20px;text-align:center;">
          <div style="font-size:2.5rem;margin-bottom:8px;">📷</div>
          <div style="font-family:'Bebas Neue',sans-serif;letter-spacing:3px;font-size:1rem;color:#9ab5cc;">CAMERA STANDBY</div>
          <div style="font-size:0.75rem;color:#6a90a8;margin-top:6px;">Upload database → Click START</div>
        </div>""",unsafe_allow_html=True)

st.markdown("<br>",unsafe_allow_html=True)
c1,c2=st.columns(2)
with c1: start=st.button("▶  START MONITORING",disabled=(not emb_ok or is_live),use_container_width=True,key="sb")
with c2: stop=st.button("■  STOP & SAVE",disabled=(not is_live),use_container_width=True,key="stb")

if not emb_ok:
    st.markdown('<div style="text-align:center;color:#4a6a88;font-size:0.68rem;letter-spacing:1px;margin-top:4px;">UPLOAD FACE DATABASE TO ENABLE</div>',unsafe_allow_html=True)

if start and emb_ok and not is_live:
    st.session_state.in_count=0; st.session_state.out_count=0
    st.session_state.events=[]; st.session_state.stop_flag=False
    st.session_state.running=True; init_excel(); st.rerun()

if stop and is_live:
    st.session_state.stop_flag=True; st.session_state.running=False; st.rerun()


# ── Load model at startup (outside running block) ─────────────
# @st.cache_resource caches across all reruns — loads only once ever
@st.cache_resource(show_spinner="⏳ Loading face recognition model...")
def load_model():
    from insightface.app import FaceAnalysis
    app=FaceAnalysis(name='buffalo_l',providers=['CUDAExecutionProvider','CPUExecutionProvider'])
    app.prepare(ctx_id=0,det_size=(160,160))
    return app

face_app=load_model()  # runs at page load, cached after first time

# ── LIVE LOOP ─────────────────────────────────────────────────
if st.session_state.running:

    names,matrix=load_embeddings(st.session_state.emb_data)
    haar=cv2.CascadeClassifier(cv2.data.haarcascades+"haarcascade_frontalface_default.xml")

    source=rtsp_url if rtsp_url else cam_index
    cap=cv2.VideoCapture(source)
    if not cap.isOpened():
        st.error(f"Cannot open camera {source}"); st.session_state.running=False; st.stop()

    cap.set(cv2.CAP_PROP_FRAME_WIDTH,640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT,480)
    cap.set(cv2.CAP_PROP_FPS,30)
    cap.set(cv2.CAP_PROP_BUFFERSIZE,1)

    tracks={}; next_tid=0; cooldown_log={}; frame_idx=0
    MAX_LOST      = 45    # ~3 seconds at 15fps — prevents false OUTs from Haar blinking
    MATCH_DIST    = 100
    STABLE_FRAMES = 6     # must be detected for 6 frames before ArcFace runs
    MIN_CONF      = 0.35  # minimum confidence to count as a known person
    MIN_FACE_AREA = 6000  # raised: ~77x77px min — rejects small false positives
    # Global identity cache: name → (roll, conf, ev_idx)
    # Once someone is identified, they are NEVER re-scanned this session
    # If they re-enter frame, we instantly restore their identity
    identity_cache = {}   # kept for compatibility
    # person_state: the real source of truth, keyed by name
    # inside, ev_idx, last_seen, out_pending_since
    person_state  = {}

    # ── Background recognition thread ─────────────────────────
    recog_queue   = queue.Queue(maxsize=2)   # small queue = always fresh crops
    result_queue  = queue.Queue()
    recog_running = [True]

    # Get the ArcFace recognition model directly (skip the detector step)
    # face_app.models has the recognition model which is much faster
    # than calling face_app.get() which re-runs detection on the crop
    rec_model = None
    for m in face_app.models.values():
        if hasattr(m,'get_feat'):
            rec_model=m; break

    def recognition_worker():
        while recog_running[0]:
            try: tid,crop = recog_queue.get(timeout=0.5)
            except queue.Empty: continue
            name,conf='Unknown',0.0
            try:
                face_112=cv2.resize(crop,(112,112))
                if rec_model is not None:
                    # Direct embedding extraction — no re-detection
                    feat=rec_model.get_feat([face_112])
                    emb=np.array(feat).flatten()
                    if emb.shape[0]>0:
                        name,conf=recognize(emb,names,matrix,threshold)
                    else:
                        raise ValueError("empty feat")
                else:
                    raise ValueError("no rec_model")
            except Exception:
                # Reliable fallback — let insightface handle it fully
                try:
                    r=face_app.get(crop)
                    if r and hasattr(r[0],'normed_embedding'):
                        name,conf=recognize(r[0].normed_embedding,names,matrix,threshold)
                except Exception: pass
            result_queue.put((tid,name,conf))

    worker=threading.Thread(target=recognition_worker,daemon=True)
    worker.start()

    st.markdown('<div class="lb">🔴 LIVE — Video stays smooth. Recognition runs in background. Appear = IN · Leave = OUT</div>',unsafe_allow_html=True)

    while st.session_state.running and not st.session_state.stop_flag:
        ret,frame=cap.read()
        if not ret: time.sleep(0.02); continue
        now=time.time(); H,W=frame.shape[:2]

        # ═══ ARCHITECTURE ════════════════════════════════════════
        # person_state: keyed by NAME, not track ID.
        # This is the source of truth for IN/OUT and re-scan prevention.
        # tracks: short-lived, just for bbox/display. Die in 0.3s.
        # When Haar sees a face → track created → ArcFace runs once →
        #   result stored in person_state[name].
        # New track of same person → ArcFace result fetched from person_state instantly.
        # ═════════════════════════════════════════════════════════
        BOX_TIMEOUT = 0.35  # seconds: box gone this fast after face leaves frame
        OUT_GRACE   = 5.0   # seconds: OUT logged only after 5s continuous absence

        # ── Collect ArcFace results ───────────────────────────
        while not result_queue.empty():
            tid,name,conf=result_queue.get_nowait()
            if tid not in tracks: continue
            roll=ROLL_NUMBERS.get(name.lower(),'—')
            tracks[tid].update({'name':name,'conf':conf,'roll':roll,
                                'identified':True,'in_queue':False})
            # Store in person_state — prevents re-scan of same person
            if name!='Unknown':
                if name not in person_state:
                    person_state[name]={'roll':roll,'conf':conf,
                                        'inside':False,'ev_idx':None,
                                        'last_seen':now,'out_pending_since':None}
                # If already inside, restore logged_in on this track
                if person_state[name]['inside']:
                    tracks[tid]['logged_in']=True
                    tracks[tid]['ev_idx']=person_state[name]['ev_idx']
                person_state[name]['last_seen']=now

        # ── Haar detection ────────────────────────────────────
        scale=320/W
        small=cv2.resize(frame,(320,int(H*scale)))
        gray=cv2.cvtColor(small,cv2.COLOR_BGR2GRAY)
        cv2.equalizeHist(gray,gray)
        # minNeighbors=12 (up from 8) — much stricter cascade, kills most
        # light/curtain false positives before any pixel checks even run
        dets=haar.detectMultiScale(gray,scaleFactor=1.05,minNeighbors=12,
                                   minSize=(55,55),flags=cv2.CASCADE_SCALE_IMAGE)
        bboxes=[]
        for (x,y,w,h) in (dets if len(dets)>0 else []):
            x1,y1=int(x/scale),int(y/scale)
            x2,y2=int((x+w)/scale),int((y+h)/scale)
            fw,fh=x2-x1,y2-y1

            # ── Filter 1: size & aspect ratio ─────────────────
            # Real faces are roughly square (0.70–1.35).
            # Curtains are tall, ceiling lights are wide — both fail this.
            if fw*fh < MIN_FACE_AREA: continue
            ar=fw/(fh+1e-5)
            if not (0.70<=ar<=1.35): continue

            # ── Filter 2: brightness — reject lights/glare ────
            # Lights appear as uniformly very bright blobs (mean > 200
            # even after histogram equalisation)
            roi_gray=gray[y:y+h,x:x+w]
            mean_bright=float(roi_gray.mean())
            if mean_bright>200: continue          # blown-out bright blob → skip

            # ── Filter 3: texture variance ─────────────────────
            # A real face has edges, shadows, features → high variance.
            # A plain curtain or light globe is nearly uniform → low variance.
            variance=float(roi_gray.var())
            if variance<150: continue             # flat/uniform region → skip

            # ── Filter 4: skin-tone hue check ─────────────────
            # Convert the full-resolution ROI to HSV and measure the
            # fraction of pixels that fall within skin-tone hue ranges.
            # Lights (white/yellow) and curtains (grey/blue/green)
            # almost never pass this threshold.
            roi_bgr=frame[y1:y2,x1:x2]
            if roi_bgr.size==0: continue
            hsv=cv2.cvtColor(roi_bgr,cv2.COLOR_BGR2HSV)
            # Two ranges cover South Asian + general skin tones
            mask1=cv2.inRange(hsv,(0,25,60),(20,180,255))
            mask2=cv2.inRange(hsv,(160,25,60),(180,180,255))
            skin_frac=float(cv2.bitwise_or(mask1,mask2).sum()/255)/(fw*fh+1e-5)
            if skin_frac<0.08: continue           # <8% skin pixels → not a face

            bboxes.append((x1,y1,x2,y2))

        # ── Match detections to existing tracks ───────────────
        matched_t=set(); matched_d=set()
        for tid in list(tracks.keys()):
            tx,ty=tracks[tid]['cx'],tracks[tid]['cy']
            bd,bi=9999,-1
            for di,(x1,y1,x2,y2) in enumerate(bboxes):
                d=np.hypot((x1+x2)//2-tx,(y1+y2)//2-ty)
                if d<MATCH_DIST and d<bd: bd=d; bi=di
            if bi>=0:
                x1,y1,x2,y2=bboxes[bi]
                cx,cy=(x1+x2)//2,(y1+y2)//2
                tracks[tid].update({'cx':cx,'cy':cy,'bbox':(x1,y1,x2,y2),
                                    'last_seen':now,'stable':tracks[tid]['stable']+1})
                # Update person_state last_seen for known persons
                nm=tracks[tid]['name']
                if nm in person_state:
                    person_state[nm]['last_seen']=now
                    person_state[nm]['out_pending_since']=None  # reset: still here
                matched_t.add(tid); matched_d.add(bi)

        # ── Expire dead tracks (BOX_TIMEOUT) ─────────────────
        # Tracks die fast so they CANNOT transfer name to new faces
        dead=[]
        for tid,tr in tracks.items():
            if tid not in matched_t and (now-tr['last_seen'])>BOX_TIMEOUT:
                dead.append(tid)
        for tid in dead: del tracks[tid]

        # ── New detections → new tracks ───────────────────────
        # Created AFTER dead tracks removed — no name inheritance possible
        for di,(x1,y1,x2,y2) in enumerate(bboxes):
            if di not in matched_d:
                tid=next_tid; next_tid+=1
                cx,cy=(x1+x2)//2,(y1+y2)//2
                # Check person_state: if we know someone and they're not
                # currently tracked, this is probably them returning
                # But we still require ArcFace confirmation — no position guessing
                tracks[tid]={'cx':cx,'cy':cy,'bbox':(x1,y1,x2,y2),
                             'last_seen':now,'stable':1,
                             'name':'...','conf':0.0,'roll':'—',
                             'identified':False,'logged_in':False,
                             'in_queue':False,'ev_idx':None}

        # ── Submit to ArcFace — skip if name already in person_state ─
        # person_state[name] means we already confirmed this face once.
        # The new track will get their identity from the result callback above.
        # Re-scan only happens if it's genuinely a new unknown face.
        for tid,tr in list(tracks.items()):
            if tr['identified'] or tr['in_queue']: continue
            if tr['stable']<STABLE_FRAMES: continue
            if now-tr['last_seen']>0.1: continue    # not detected this frame
            x1,y1,x2,y2=tr['bbox']
            pad=20
            crop=frame[max(0,y1-pad):min(H,y2+pad),max(0,x1-pad):min(W,x2+pad)].copy()
            if crop.size==0: tracks[tid]['identified']=True; continue
            if not recog_queue.full():
                recog_queue.put((tid,crop)); tracks[tid]['in_queue']=True

        # ── Update person_state out_pending clock ─────────────
        # For known persons not matched this frame, start/continue out clock
        active_names={tr['name'] for tr in tracks.values()
                      if now-tr['last_seen']<0.1 and tr['name'] not in('...','Unknown')}
        for name,ps in person_state.items():
            if name in active_names:
                ps['out_pending_since']=None  # still visible — reset
            else:
                if ps['inside'] and ps['out_pending_since'] is None:
                    ps['out_pending_since']=now  # start counting absence

        # ── IN events — enrolled students only ──────────────
        # Unknown = curtain / light / object → ignored completely
        for tid,tr in list(tracks.items()):
            if now-tr['last_seen']>0.1: continue
            name=tr['name']; roll=tr['roll']; conf=tr['conf']
            # Only log if: ArcFace confirmed + name is in our database + conf high enough
            if not tr['identified']: continue
            if name in ('...','Unknown'): continue
            if conf < MIN_CONF: continue
            if tr['logged_in']: continue
            # Extra guard: name must be in ROLL_NUMBERS (enrolled student)
            if name.lower() not in ROLL_NUMBERS: continue
            log_key=name
            if now-cooldown_log.get(log_key,0)<cooldown:
                tracks[tid]['logged_in']=True; continue
            cooldown_log[log_key]=now
            st.session_state.in_count+=1
            t=datetime.now().strftime('%H:%M:%S')
            ev={'name':name,'roll':roll,'in_time':t,'out_time':'—',
                'date':datetime.now().strftime('%d-%m-%Y')}
            st.session_state.events.append(ev)
            ev_idx=len(st.session_state.events)-1
            tracks[tid]['ev_idx']=ev_idx; tracks[tid]['logged_in']=True
            person_state[name]['inside']=True
            person_state[name]['ev_idx']=ev_idx
            person_state[name]['out_pending_since']=None
            log_excel(name,roll,'IN',conf)

        # ── OUT events — enrolled students only ────────────
        # Fires only after OUT_GRACE seconds of confirmed absence.
        # Unknown/objects never reach person_state so never fire OUT.
        for name,ps in list(person_state.items()):
            if not ps['inside']: continue
            if ps['out_pending_since'] is None: continue
            if now-ps['out_pending_since'] < OUT_GRACE: continue
            if name.lower() not in ROLL_NUMBERS: continue  # enrolled only
            roll=ps['roll']; conf=ps['conf']
            okey=name+'_out'
            if now-cooldown_log.get(okey,0)<cooldown: continue
            cooldown_log[okey]=now
            st.session_state.out_count+=1
            t=datetime.now().strftime('%H:%M:%S')
            ev_idx=ps['ev_idx']
            if ev_idx is not None and ev_idx<len(st.session_state.events):
                st.session_state.events[ev_idx]['out_time']=t
            ps['inside']=False; ps['ev_idx']=None; ps['out_pending_since']=None
            log_excel(name,roll,'OUT',conf)

        # ── Draw — only confirmed/identifying tracks ───────────
        # Tracks that are still in the "stable" phase (not yet sent to
        # ArcFace) are intentionally NOT drawn — this prevents the
        # flickering "Scanning..." boxes that appear over lights/curtains.
        for tid,tr in tracks.items():
            if now-tr['last_seen']>BOX_TIMEOUT: continue
            # Skip if not yet stable enough to have been sent to ArcFace
            if tr['stable'] < STABLE_FRAMES and not tr['identified']: continue
            x1,y1,x2,y2=tr['bbox']; name=tr['name']; roll=tr['roll']
            pending=name=='...'
            unknown=name=='Unknown'
            known=not pending and not unknown
            # Color: blue=identifying, green=enrolled, hide unknown entirely
            if unknown: continue                  # don't draw boxes for non-faces
            color=(0,180,255) if pending else (0,200,100)
            cv2.rectangle(frame,(x1,y1),(x2,y2),color,2)
            if pending:
                label="Identifying..."
            else:
                label=name.title()
            (tw,th),_=cv2.getTextSize(label,cv2.FONT_HERSHEY_SIMPLEX,0.55,1)
            cv2.rectangle(frame,(x1,y1-th-20),(x1+max(tw,60)+8,y1),color,-1)
            cv2.putText(frame,label,(x1+3,y1-12),cv2.FONT_HERSHEY_SIMPLEX,0.55,(0,0,0),1)
            if known and roll!='—':
                cv2.putText(frame,roll,(x1+3,y1-1),cv2.FONT_HERSHEY_SIMPLEX,0.38,(0,0,0),1)


        inside=max(0,st.session_state.in_count-st.session_state.out_count)
        ov=frame.copy(); cv2.rectangle(ov,(0,0),(200,94),(4,8,18),-1)
        cv2.addWeighted(ov,0.72,frame,0.28,0,frame)
        cv2.putText(frame,f"IN    : {st.session_state.in_count}",(8,25),cv2.FONT_HERSHEY_SIMPLEX,0.65,(0,255,100),2)
        cv2.putText(frame,f"OUT   : {st.session_state.out_count}",(8,52),cv2.FONT_HERSHEY_SIMPLEX,0.65,(0,80,255),2)
        cv2.putText(frame,f"INSIDE: {inside}",(8,79),cv2.FONT_HERSHEY_SIMPLEX,0.55,(255,220,0),2)

        try: feed_ph.image(cv2.cvtColor(frame,cv2.COLOR_BGR2RGB),channels="RGB",width="stretch")
        except Exception: feed_ph.image(cv2.cvtColor(frame,cv2.COLOR_BGR2RGB),channels="RGB",use_container_width=True)

    recog_running[0]=False   # signal background thread to stop
    cap.release()
    if not st.session_state.running: st.rerun()


# ── AFTER STOP ───────────────────────────────────────────────
if not st.session_state.running and st.session_state.events:
    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<div class="sh"><h3>📊 SESSION SUMMARY</h3></div>',unsafe_allow_html=True)
    inside=max(0,st.session_state.in_count-st.session_state.out_count)
    st.markdown(f"""<div class="mrow">
      <div class="mc in"><div class="ml">Total Entries</div><div class="mv">{st.session_state.in_count}</div><div class="ms">persons entered</div></div>
      <div class="mc out"><div class="ml">Total Exits</div><div class="mv">{st.session_state.out_count}</div><div class="ms">persons exited</div></div>
      <div class="mc now"><div class="ml">Currently Inside</div><div class="mv">{inside}</div><div class="ms">occupancy</div></div>
    </div>""",unsafe_allow_html=True)

    rows=""
    for ev in st.session_state.events:
        nm=ev['name'].title() if ev['name']!='Unknown' else '<span style="color:#6080a0">Unknown</span>'
        out_disp=(f'<span style="color:#ef4444;font-family:JetBrains Mono,monospace;font-size:0.72rem;">{ev["out_time"]}</span>'
                  if ev['out_time']!='—'
                  else '<span style="color:#f59e0b;font-size:0.68rem;">still inside</span>')
        rows+=(f'<div class="er" style="grid-template-columns:1.4fr 1.2fr 1fr 1fr;">'
               f'<div style="font-weight:500;">{nm}</div>'
               f'<div style="font-family:JetBrains Mono,monospace;font-size:0.7rem;color:#8baec8;">{ev["roll"]}</div>'
               f'<div style="color:#10b981;font-family:JetBrains Mono,monospace;font-size:0.72rem;">{ev["in_time"]}</div>'
               f'<div>{out_disp}</div>'
               f'</div>')
    st.markdown(f'<div class="el"><div class="eh" style="grid-template-columns:1.4fr 1.2fr 1fr 1fr;"><span>Name</span><span>Roll No.</span><span>Time In</span><span>Time Out</span></div>{rows}</div>',unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<div class="sh"><h3>⬇ DOWNLOAD</h3></div>',unsafe_allow_html=True)
    dc1,dc2=st.columns(2)
    with dc1:
        sf=st.session_state.get('session_file')
        if sf and os.path.exists(sf):
            with open(sf,'rb') as f: sb=f.read()
            st.download_button("📋  This Session Only",data=sb,
                file_name=sf,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
    with dc2:
        if os.path.exists(EXCEL_MASTER):
            with open(EXCEL_MASTER,'rb') as f: mb=f.read()
            st.download_button("📊  All Sessions (Master)",data=mb,
                file_name="attendance_master.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)